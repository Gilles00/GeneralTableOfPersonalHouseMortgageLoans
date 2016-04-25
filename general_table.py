# -*- encoding:utf-8 -*-
import MySQLdb
import numpy as np
import pandas as pd
import sys

try:
    conn = MySQLdb.connect(
            host = '10.0.9.127',
            port = 3306,
            user = 'alas',
            passwd = '6143',
            db = 'data_personal_house_mortgage_loans',
            charset = 'utf8',)
except Exception, e:
    print e
    sys.exit()
cur = conn.cursor()

month = ['january', 'february', 'march', 'april', 'may', 'june', 
         'july', 'august', 'september', 'october', 'november', 'december']

#city = [u'北京', u'上海', u'广州',  u'深圳']

city = [u'北京', u'上海', u'广州', u'深圳', u'杭州', u'南京', u'武汉',
        u'重庆', u'成都', u'天津', u'青岛', u'长沙', u'苏州', u'厦门',
        u'郑州', u'西安', u'无锡', u'哈尔滨', u'佛山', u'珠海', u'东莞',
        u'海口', u'宁波', u'合肥', u'南昌', u'大连', u'昆明', u'济南', 
        u'沈阳', u'太原', u'乌鲁木齐', u'南宁', u'福州', u'长春', u'石家庄',]

ave_loan_rates_diff_city = np.zeros((3, 35, 2))
prop_discount_first_home = np.zeros((3, 35, 1))
prop_stop_handling_loans = np.zeros((3, 35, 1))
ave_loan_rates_first_home_diff_bank = np.zeros((3, 19, 1))
prop_ave_loan_rates_first_home_diff_city = np.zeros((3, 35, 4))
prop_ave_loan_rates_second_home_diff_city = np.zeros((3, 35, 4))
prop_ave_loan_rates_1and2_home = np.zeros((3, 2, 4))

for i in range(len(month)):
    month[i] += '2015'
    for j in range(len(city)):
        mysql = """
            select
                loan_rates_first_home,
                loan_rates_second_home
            from %s
            where city="%s"
                """ % (month[i], city[j])
        try:
            df = pd.read_sql(mysql, con=conn)
        except Exception, e:
            print e
        #将首套二套的利率值由unicode (举个栗子u'0.9')转为float (0.9)
        df[['loan_rates_first_home', 'loan_rates_second_home']] = df[
                ['loan_rates_first_home', 'loan_rates_second_home']].astype(float)
        #停贷的不参与均值计算
        df[df>500] = np.nan
        
        #各城市平均利率
        ave_loan_rates_diff_city[i, j] = np.around(
                df[['loan_rates_first_home', 
                    'loan_rates_second_home']].mean()*4.9,
                decimals = 3)
#        print month[i], city[j], '\t\t', ave_loan_rates_diff_city[i, j]
#        print '\n'

        #优惠占比
        df2 = df['loan_rates_first_home']
#        print df2[df2<1].count(), df2[df2>=1].count(), df2.count(), len(df2)
        prop_discount_first_home[i, j] = np.around(
                1.0*df2[df2<1].count()/len(df2),  decimals = 4)
#        print month[i], city[j], '\t\t', prop_discount_first_home[i, j]

        #停贷占比
        prop_stop_handling_loans[i, j] = np.around(
                1.0*(len(df2)-df2.count())/len(df2), decimals = 4)
#        print month[i], city[j], prop_stop_handling_loans[i, j]
#    print '\n'
        #输出各城市停贷银行数
#        if i == 2:
#            print len(df2)-df2.count()

        #首套房利率分类占比
        prop_ave_loan_rates_first_home_diff_city[i, j, 0] = np.around(
            1.0*df2[df2 <0.9].count()/len(df2), decimals = 4) 
        prop_ave_loan_rates_first_home_diff_city[i, j, 1] = np.around(
            1.0*df2[0.9<= df2][df2[0.9<= df2] <1].count()/len(df2), decimals = 4) 
        prop_ave_loan_rates_first_home_diff_city[i, j, 2] = np.around(
            1.0*df2[1<= df2].count()/len(df2), decimals = 4) 
        prop_ave_loan_rates_first_home_diff_city[i, j, 3] = np.around(
            1.0*(len(df2)-df2.count())/len(df2), decimals = 4)
#        print month[i], city[j], prop_ave_loan_rates_first_home_diff_city[i, j]
#        print 1.0*df2[df2 <0.9].count(),
#        print 1.0*df2[0.9<= df2][df2[0.9<= df2] <1].count(),
#        print 1.0*df2[1<= df2].count(),
#        print 1.0*(len(df2)-df2.count()), '\n'
#    print '\n'

        #二套房利率分类占比
        df3 = df['loan_rates_second_home']
        prop_ave_loan_rates_second_home_diff_city[i, j, 0] = np.around(
            1.0*df3[df3 <=1].count()/len(df3), decimals = 4) 
        prop_ave_loan_rates_second_home_diff_city[i, j, 1] = np.around(
            1.0*df3[1< df3][df3[1< df3] <=1.1].count()/len(df3), decimals = 4) 
        prop_ave_loan_rates_second_home_diff_city[i, j, 2] = np.around(
            1.0*df3[1.1< df3].count()/len(df3), decimals = 4) 
        prop_ave_loan_rates_second_home_diff_city[i, j, 3] = np.around(
            1.0*(len(df3)-df3.count())/len(df3), decimals = 4)
#        print month[i], city[j], prop_ave_loan_rates_second_home_diff_city[i, j],
#        print 1.0*df3[df3 <=1].count(),
#        print 1.0*df3[1< df3][df3[1< df3] <=1.1].count(),
#        print 1.0*df3[1.1< df3].count(),
#        print 1.0*(len(df3)-df3.count()), '\n'
#    print '\n'

#全国首套和二套利率的分类占比 
for i in range(len(month)):
    mysql2 = """
            select
                loan_rates_first_home,
                loan_rates_second_home
            from %s
                """ % month[i]
    try:
        df = pd.read_sql(mysql2, con=conn)
    except Exception, e:
        print e
    df[['loan_rates_first_home', 'loan_rates_second_home']] = df[
        ['loan_rates_first_home', 'loan_rates_second_home']].astype(float)
    df[df>500] = np.nan

    df2 = df['loan_rates_first_home']
    prop_ave_loan_rates_1and2_home[i, 0, 0] = np.around(
        1.0*df2[df2 <0.9].count()/len(df2), decimals = 4) 
    prop_ave_loan_rates_1and2_home[i, 0, 1] = np.around(
        1.0*df2[0.9<= df2][df2[0.9<= df2] <1].count()/len(df2), decimals = 4) 
    prop_ave_loan_rates_1and2_home[i, 0, 2] = np.around(
        1.0*df2[1<= df2].count()/len(df2), decimals = 4) 
    prop_ave_loan_rates_1and2_home[i, 0, 3] = np.around(
        1.0*(len(df2)-df2.count())/len(df2), decimals = 4)

    df3 = df['loan_rates_second_home']
    prop_ave_loan_rates_1and2_home[i, 1, 0] = np.around(
        1.0*df3[df3 <=1].count()/len(df3), decimals = 4) 
    prop_ave_loan_rates_1and2_home[i, 1, 1] = np.around(
        1.0*df3[1< df3][df3[1< df3] <=1.1].count()/len(df3), decimals = 4) 
    prop_ave_loan_rates_1and2_home[i, 1, 2] = np.around(
        1.0*df3[1.1< df3].count()/len(df3), decimals = 4) 
    prop_ave_loan_rates_1and2_home[i, 1, 3] = np.around(
        1.0*(len(df3)-df3.count())/len(df3), decimals = 4)
'''
print prop_ave_loan_rates_1and2_home
print 1.0*df2[df2 <0.9].count(),
print 1.0*df2[0.9<= df2][df2[0.9<= df2] <1].count(),
print 1.0*df2[1<= df2].count(),
print 1.0*(len(df2)-df2.count()), '\n'
print len(df2)

print 1.0*df3[df3 <=1].count(),
print 1.0*df3[1< df3][df3[1< df3] <=1.1].count(),
print 1.0*df3[1.1< df3].count(),
print 1.0*(len(df3)-df3.count()), '\n'
print len(df3)
'''
#bank = [u'工商银行', u'农业银行']
bank = [u'工商银行', u'农业银行', u'中国银行', u'建设银行', u'交通银行', u'邮储银行',
        u'中信银行', u'光大银行', u'民生银行', u'招商银行', u'广发银行', u'浦发银行',
        u'兴业银行', u'华夏银行', u'东亚银行', u'渣打银行', u'花旗银行', u'汇丰银行',
        u'恒生银行']
for i in range(len(month)):
    for j in range(len(bank)):
        mysql3 = """
            select
                loan_rates_first_home
            from %s
            where bank = "%s"
                """ % (month[i], bank[j])
        try:
            df = pd.read_sql(mysql3, con=conn)
        except Exception, e:
            print e
        df['loan_rates_first_home'] = df['loan_rates_first_home'].astype(float)
        df[df>500] = np.nan
        df4 = df['loan_rates_first_home']
        ave_loan_rates_first_home_diff_bank[i, j, 0] = np.around(
                df4.mean(), decimals = 4)
#        print ave_loan_rates_first_home_diff_bank[i, j, 0],
#    print '\n'
    

from openpyxl import Workbook

wb = Workbook()
ws1 = wb.active
ws1.title = u'35城市首套房平均利率'
ws1.append([''] + month) 
for j in range(len(city)):
    l = []
    l.append(city[j])
    for i in range(len(month)):
        l.append(ave_loan_rates_diff_city[i, j, 0])
    ws1.append(l) 

ws2 = wb.create_sheet()
ws2.title = u'35城市二套房平均利率'
ws2.append([''] + month) 
for j in range(len(city)):
    l = []
    l.append(city[j])
    for i in range(len(month)):
        l.append(ave_loan_rates_diff_city[i, j, 1])
    ws2.append(l) 

ws23 = wb.create_sheet()
ws23.title = u'各城市每月平均利率环比升降'
ws23.append([u'银行'] + month[::-1][:2] + [u'两月利率较小者', u'同比增加按百分点数', u'同比降低百分点数'])

for j in range(len(bank)):
    l = []
    l.append(bank[j])
    for i in range(len(month)-1):
        l.append(ave_loan_rates_first_home_diff_bank[len(month)-1-i, j, 0])
    a = ave_loan_rates_first_home_diff_bank[len(month)-1, j, 0]
    b = ave_loan_rates_first_home_diff_bank[len(month)-2, j, 0] 
    if a > b:
        l.append(b)
        l.append(np.around(a-b, decimals = 3)) 
        l.append(0)
    else:
        l.append(a)
        l.append(0)
        l.append(np.around(b-a, decimals = 3))
    ws23.append(l)

ws3 = wb.create_sheet()
ws3.title = u'优惠占比'
ws3.append([''] + month)
for j in range(len(city)):
    l = []
    l.append(city[j])
    for i in range(len(month)):
        l.append(prop_discount_first_home[i, j, 0])
    ws3.append(l)

ws4 = wb.create_sheet()
ws4.title = u'停贷占比'
ws4.append([''] + month) 
for j in range(len(city)):
    l = []
    l.append(city[j])
    for i in range(len(month)):
        l.append(prop_stop_handling_loans[i, j, 0])
    ws4.append(l)

ws5 = wb.create_sheet()
ws5.title = u'首套房利率分类占比'
ws5.merge_cells('B1:E1')
ws5['B1'] = month[0]
ws5.merge_cells('F1:I1')
ws5['F1'] = month[1]
ws5.merge_cells('J1:M1')
ws5['J1'] = month[2]
prop_clas_first_home = [
        '9折以下', '9折(含)~基准', '基准及以上', '停贷']
ws5.append([''] + prop_clas_first_home*len(month)) 

l = [u'全国']
prop_ave_loan_rates_1and2_home
for i in range(len(month)):
    for x in range(4):
        l.append(prop_ave_loan_rates_1and2_home[i, 0, x])
ws5.append(l)

for j in range(len(city)):
    l = []
    l.append(city[j])
    for i in range(len(month)):
        for f in range(4):
            l.append(prop_ave_loan_rates_first_home_diff_city[i, j, f])
    ws5.append(l)

ws6 = wb.create_sheet()
ws6.title = u'二套房利率分类占比'
ws6.merge_cells('B1:E1')
ws6['B1'] = month[0]
ws6.merge_cells('F1:I1')
ws6['F1'] = month[1]
ws6.merge_cells('J1:M1')
ws6['J1'] = month[2]
prop_clas_second_home = [
        '基准及以下', '基准~基准上浮10%(含)', '基准上浮以上', '停贷']
ws6.append([''] + prop_clas_second_home*len(month)) 

l = [u'全国']
prop_ave_loan_rates_1and2_home
for i in range(len(month)):
    for x in range(4):
        l.append(prop_ave_loan_rates_1and2_home[i, 1, x])
ws6.append(l)

for j in range(len(city)):
    l = []
    l.append(city[j])
    for i in range(len(month)):
        for f in range(4):
            l.append(prop_ave_loan_rates_second_home_diff_city[i, j, f])
    ws6.append(l)


wb.save("/users/alas/desktop/general_table.xlsx")

